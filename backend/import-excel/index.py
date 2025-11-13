'''
Business: Импорт данных о военнослужащих из Excel файлов в базу данных
Args: event с httpMethod, body (base64 encoded Excel file)
Returns: HTTP response с результатами импорта
'''

import json
import base64
import os
from typing import Dict, Any, List, Optional
from datetime import datetime
import psycopg2
from psycopg2.extras import RealDictCursor
from openpyxl import load_workbook
from io import BytesIO

def parse_date(value: Any) -> Optional[str]:
    if not value:
        return None
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d')
    if isinstance(value, str):
        try:
            for fmt in ['%d.%m.%Y', '%Y-%m-%d', '%d/%m/%Y']:
                try:
                    return datetime.strptime(value.strip(), fmt).strftime('%Y-%m-%d')
                except ValueError:
                    continue
        except:
            pass
    return None

def parse_rank(rank_str: str) -> Optional[str]:
    if not rank_str:
        return None
    rank_map = {
        'рядовой': 'private', 'рядовий': 'private',
        'ефрейтор': 'corporal',
        'младший сержант': 'junior_sergeant', 'мл сержант': 'junior_sergeant',
        'сержант': 'sergeant',
        'старший сержант': 'senior_sergeant', 'ст сержант': 'senior_sergeant',
        'старшина': 'foreman',
        'прапорщик': 'warrant_officer',
        'старший прапорщик': 'senior_warrant_officer', 'ст прапорщик': 'senior_warrant_officer',
        'младший лейтенант': 'junior_lieutenant', 'мл лейтенант': 'junior_lieutenant',
        'лейтенант': 'lieutenant',
        'старший лейтенант': 'senior_lieutenant', 'ст лейтенант': 'senior_lieutenant',
        'капитан': 'captain',
        'майор': 'major',
        'подполковник': 'lieutenant_colonel',
        'полковник': 'colonel'
    }
    return rank_map.get(rank_str.lower().strip())

def parse_status(status_str: str) -> str:
    if not status_str:
        return 'active'
    status_map = {
        'находится': 'active', 'в части': 'active', 'активный': 'active',
        'отпуск': 'leave', 'отпуска': 'leave',
        'командировка': 'business_trip', 'в командировке': 'business_trip',
        'госпитализация': 'hospitalized', 'госпиталь': 'hospitalized', 'в госпитале': 'hospitalized',
        'вкк': 'vkk', 'ввк': 'vvk', 'цввк': 'cvvk',
        'пвд': 'pvd',
        'ввк на изменение категории': 'vvk_category_change',
        'амбулаторное лечение': 'ambulatory_treatment', 'амбулаторное': 'ambulatory_treatment',
        'увольнение': 'discharge', 'уволен': 'discharge'
    }
    return status_map.get(status_str.lower().strip(), 'active')

def parse_fitness_category(cat_str: str) -> Optional[str]:
    if not cat_str:
        return None
    cat_str = cat_str.strip().upper()
    if cat_str in ['А', 'A', 'А1', 'А2', 'А3', 'А4']:
        return 'A'
    if cat_str in ['Б', 'B', 'Б1', 'Б2', 'Б3', 'Б4']:
        return 'B'
    if cat_str in ['В', 'V', 'В1', 'В2', 'В3', 'В4']:
        return 'V'
    if cat_str in ['Г', 'G']:
        return 'G'
    if cat_str in ['Д', 'D']:
        return 'D'
    return None

def handler(event: Dict[str, Any], context: Any) -> Dict[str, Any]:
    method: str = event.get('httpMethod', 'POST')
    
    if method == 'OPTIONS':
        return {
            'statusCode': 200,
            'headers': {
                'Access-Control-Allow-Origin': '*',
                'Access-Control-Allow-Methods': 'POST, OPTIONS',
                'Access-Control-Allow-Headers': 'Content-Type, X-User-Id',
                'Access-Control-Max-Age': '86400'
            },
            'body': ''
        }
    
    if method != 'POST':
        return {
            'statusCode': 405,
            'headers': {
                'Content-Type': 'application/json',
                'Access-Control-Allow-Origin': '*'
            },
            'isBase64Encoded': False,
            'body': json.dumps({'error': 'Method not allowed'}, ensure_ascii=False)
        }
    
    try:
        body_str = event.get('body', '{}')
        if not body_str or body_str.strip() == '':
            body_str = '{}'
        
        try:
            body_data = json.loads(body_str)
        except json.JSONDecodeError:
            return {
                'statusCode': 400,
                'headers': {
                    'Content-Type': 'application/json',
                    'Access-Control-Allow-Origin': '*'
                },
                'isBase64Encoded': False,
                'body': json.dumps({'error': 'Невалидный JSON'}, ensure_ascii=False)
            }
        
        file_base64 = body_data.get('file')
        
        if not file_base64:
            return {
                'statusCode': 400,
                'headers': {
                    'Content-Type': 'application/json',
                    'Access-Control-Allow-Origin': '*'
                },
                'isBase64Encoded': False,
                'body': json.dumps({'error': 'Файл не предоставлен'}, ensure_ascii=False)
            }
        
        file_bytes = base64.b64decode(file_base64)
        wb = load_workbook(BytesIO(file_bytes), data_only=True)
        
        database_url = os.environ.get('DATABASE_URL')
        if not database_url:
            raise Exception('DATABASE_URL не установлен')
        
        conn = psycopg2.connect(database_url)
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        total_imported = 0
        all_errors = []
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet_type = 'main'
            
            if 'отпуск' in sheet_name.lower():
                sheet_type = 'leave'
            elif 'госпитал' in sheet_name.lower():
                sheet_type = 'hospitalized'
            elif 'отправ' in sheet_name.lower() or 'пвд' in sheet_name.lower():
                sheet_type = 'dispatch'
        
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return {
                'statusCode': 400,
                'headers': {
                    'Content-Type': 'application/json',
                    'Access-Control-Allow-Origin': '*'
                },
                'isBase64Encoded': False,
                'body': json.dumps({'error': 'Файл пустой или содержит только заголовки'}, ensure_ascii=False)
            }
        
        headers = [str(h).strip().lower() if h else '' for h in rows[0]]
        
        col_map = {}
        for i, h in enumerate(headers):
            if 'фио' in h or 'фамилия' in h or 'имя' in h:
                col_map['full_name'] = i
            elif 'личный номер' in h or 'личн' in h:
                col_map['personal_number'] = i
            elif 'подразделение' in h:
                col_map['unit'] = i
            elif 'звание' in h or 'ранг' in h:
                col_map['rank'] = i
            elif 'дата рождения' in h or 'др' in h or 'рождения' in h:
                col_map['birth_date'] = i
            elif 'дата прибытия' in h or 'прибытие' in h:
                col_map['arrival_date'] = i
            elif 'срок исключения' in h or 'срок окончания' in h:
                col_map['exclusion_date'] = i
            elif 'категория' in h and ('годн' in h or 'вкк' in h):
                col_map['fitness_category'] = i
            elif 'привлечение' in h:
                col_map['involvement'] = i
            elif 'склад' in h and 'ивд' in h:
                col_map['pvd_storage'] = i
            elif 'количест' in h and 'ивд' in h:
                col_map['pvd_count'] = i
            elif 'военный билет' in h or 'вб' in h or 'номер вб' in h:
                col_map['military_id'] = i
            elif 'исключение' in h and 'должности' in h:
                col_map['exclusion_reason'] = i
            elif 'проблем' in h and 'решени' in h:
                col_map['issue_status'] = i
            elif 'статус' in h or 'положение' in h:
                col_map['status'] = i
            elif 'вмо' in h:
                col_map['vmo'] = i
            elif 'диагноз' in h:
                col_map['diagnosis'] = i
            elif 'комментарий' in h or 'примечание' in h or 'заметки' in h or 'ссылка' in h:
                col_map['notes'] = i
        
            imported = 0
            errors = []
        
        for idx, row in enumerate(rows[1:], start=2):
            try:
                personal_number = str(row[col_map.get('personal_number', -1)]).strip() if col_map.get('personal_number') is not None and row[col_map.get('personal_number', -1)] else None
                
                if not personal_number or personal_number == '':
                    continue
                
                full_name = str(row[col_map.get('full_name', 0)]).strip() if col_map.get('full_name') is not None and row[col_map.get('full_name', 0)] else None
                if not full_name:
                    continue
                
                unit = str(row[col_map.get('unit', -1)]).strip() if col_map.get('unit') is not None and row[col_map.get('unit', -1)] else None
                rank = parse_rank(str(row[col_map.get('rank', -1)])) if col_map.get('rank') is not None and row[col_map.get('rank', -1)] else None
                birth_date = parse_date(row[col_map.get('birth_date', -1)]) if col_map.get('birth_date') is not None else None
                arrival_date = parse_date(row[col_map.get('arrival_date', -1)]) if col_map.get('arrival_date') is not None else None
                exclusion_date = parse_date(row[col_map.get('exclusion_date', -1)]) if col_map.get('exclusion_date') is not None else None
                fitness_category = parse_fitness_category(str(row[col_map.get('fitness_category', -1)])) if col_map.get('fitness_category') is not None and row[col_map.get('fitness_category', -1)] else None
                involvement = str(row[col_map.get('involvement', -1)]).strip() if col_map.get('involvement') is not None and row[col_map.get('involvement', -1)] else None
                pvd_storage = str(row[col_map.get('pvd_storage', -1)]).strip() if col_map.get('pvd_storage') is not None and row[col_map.get('pvd_storage', -1)] else None
                pvd_count = str(row[col_map.get('pvd_count', -1)]).strip() if col_map.get('pvd_count') is not None and row[col_map.get('pvd_count', -1)] else None
                military_id = str(row[col_map.get('military_id', -1)]).strip() if col_map.get('military_id') is not None and row[col_map.get('military_id', -1)] else None
                exclusion_reason = str(row[col_map.get('exclusion_reason', -1)]).strip() if col_map.get('exclusion_reason') is not None and row[col_map.get('exclusion_reason', -1)] else None
                issue_status = str(row[col_map.get('issue_status', -1)]).strip() if col_map.get('issue_status') is not None and row[col_map.get('issue_status', -1)] else None
                status = parse_status(str(row[col_map.get('status', -1)])) if col_map.get('status') is not None and row[col_map.get('status', -1)] else 'active'
                vmo = str(row[col_map.get('vmo', -1)]).strip() if col_map.get('vmo') is not None and row[col_map.get('vmo', -1)] else None
                diagnosis = str(row[col_map.get('diagnosis', -1)]).strip() if col_map.get('diagnosis') is not None and row[col_map.get('diagnosis', -1)] else None
                notes = str(row[col_map.get('notes', -1)]).strip() if col_map.get('notes') is not None and row[col_map.get('notes', -1)] else None
                
                cur.execute(
                    "SELECT id FROM personnel WHERE personal_number = %s",
                    (personal_number,)
                )
                existing = cur.fetchone()
                
                if existing:
                    person_id = existing['id']
                    cur.execute(
                        """
                        UPDATE personnel 
                        SET full_name = %s, rank = %s, birth_date = %s, military_id = %s, 
                            unit = %s, current_status = %s, fitness_category = %s, 
                            notes = %s, updated_at = NOW()
                        WHERE id = %s
                        """,
                        (full_name, rank, birth_date, military_id, unit, status, fitness_category, notes, person_id)
                    )
                else:
                    cur.execute(
                        """
                        INSERT INTO personnel (full_name, personal_number, rank, birth_date, military_id, unit, current_status, fitness_category, notes, created_at)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, NOW())
                        RETURNING id
                        """,
                        (full_name, personal_number, rank, birth_date, military_id, unit, status, fitness_category, notes)
                    )
                    person_id = cur.fetchone()['id']
                
                if arrival_date:
                    cur.execute(
                        """
                        INSERT INTO movements (personnel_id, movement_type, start_date, notes, created_at)
                        VALUES (%s, 'arrival', %s, %s, NOW())
                        """,
                        (person_id, arrival_date, 'Прибытие из Excel импорта')
                    )
                
                if status != 'active':
                    movement_type = status
                    if status in ['vkk', 'vvk', 'cvvk', 'vvk_category_change', 'ambulatory_treatment']:
                        movement_type = status
                    elif status == 'hospitalized':
                        movement_type = 'hospitalized'
                    elif status == 'leave':
                        movement_type = 'leave'
                    elif status == 'pvd':
                        movement_type = 'pvd'
                    
                    cur.execute(
                        """
                        INSERT INTO movements (personnel_id, movement_type, start_date, vmo, notes, created_at)
                        VALUES (%s, %s, CURRENT_DATE, %s, %s, NOW())
                        """,
                        (person_id, movement_type, vmo, notes)
                    )
                
                if fitness_category:
                    cur.execute(
                        """
                        INSERT INTO medical_checkups (personnel_id, checkup_date, diagnosis, fitness_category, notes, created_at)
                        VALUES (%s, CURRENT_DATE, %s, %s, %s, NOW())
                        """,
                        (person_id, diagnosis or 'Импорт из Excel', fitness_category, notes)
                    )
                
                if pvd_storage or pvd_count:
                    pvd_notes = f'Склад ИВД: {pvd_storage}, Количество: {pvd_count}' if pvd_storage and pvd_count else pvd_storage or pvd_count
                    if notes:
                        notes = f'{notes}. ИВД: {pvd_notes}'
                    else:
                        notes = f'ИВД: {pvd_notes}'
                    cur.execute(
                        "UPDATE personnel SET notes = %s WHERE id = %s",
                        (notes, person_id)
                    )
                
                if sheet_type == 'leave':
                    cur.execute(
                        "SELECT id FROM movements WHERE personnel_id = %s AND movement_type = 'leave' AND end_date IS NULL",
                        (person_id,)
                    )
                    if not cur.fetchone():
                        cur.execute(
                            """
                            INSERT INTO movements (personnel_id, movement_type, start_date, notes, created_at)
                            VALUES (%s, 'leave', CURRENT_DATE, %s, NOW())
                            """,
                            (person_id, notes or f'Отпуск ({sheet_name})')
                        )
                elif sheet_type == 'hospitalized':
                    cur.execute(
                        "SELECT id FROM movements WHERE personnel_id = %s AND movement_type = 'hospitalized' AND end_date IS NULL",
                        (person_id,)
                    )
                    if not cur.fetchone():
                        cur.execute(
                            """
                            INSERT INTO movements (personnel_id, movement_type, start_date, vmo, notes, created_at)
                            VALUES (%s, 'hospitalized', CURRENT_DATE, %s, %s, NOW())
                            """,
                            (person_id, vmo, notes or f'Госпитализация ({sheet_name})')
                        )
                elif sheet_type == 'dispatch':
                    cur.execute(
                        "SELECT id FROM movements WHERE personnel_id = %s AND movement_type = 'pvd' AND end_date IS NULL",
                        (person_id,)
                    )
                    if not cur.fetchone():
                        cur.execute(
                            """
                            INSERT INTO movements (personnel_id, movement_type, start_date, notes, created_at)
                            VALUES (%s, 'pvd', CURRENT_DATE, %s, NOW())
                            """,
                            (person_id, notes or f'ПВД ({sheet_name})')
                        )
                
                imported += 1
                
            except Exception as e:
                errors.append(f'[{sheet_name}] Строка {idx}: {str(e)}')
        
            total_imported += imported
            all_errors.extend(errors)
        
        conn.commit()
        cur.close()
        conn.close()
        
        return {
            'statusCode': 200,
            'headers': {
                'Content-Type': 'application/json',
                'Access-Control-Allow-Origin': '*'
            },
            'isBase64Encoded': False,
            'body': json.dumps({
                'success': True,
                'imported': total_imported,
                'errors': all_errors,
                'sheets_processed': len(wb.sheetnames)
            }, ensure_ascii=False)
        }
        
    except Exception as e:
        return {
            'statusCode': 500,
            'headers': {
                'Content-Type': 'application/json',
                'Access-Control-Allow-Origin': '*'
            },
            'isBase64Encoded': False,
            'body': json.dumps({'error': str(e)})
        }